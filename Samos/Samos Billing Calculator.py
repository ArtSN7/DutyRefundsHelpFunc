# -*- coding: utf-8 -*-
"""
Скрипт для расчета счета Duty Refunds для клиента Samos/Decantbuy
Автор: Duty Refunds
Дата: Ноябрь 2025

Описание:
Этот скрипт анализирует данные о посылках с парфюмерией, импортированных в UK,
и рассчитывает счет на основе следующей формулы:

СЧЕТ = [+] UK VAT (для UK заказов)
       [-] UK VAT (для возвратов UK заказов)  
       [+] UK Duty (обычно 0 для парфюмерии)
       [+] Комиссия: (£1 + 5% * (UK Duty + VAT)) * количество UK/EU заказов

Логика расчета:
- UK заказы (Сценарий A): Country = 'GB' и пустой UK Export Date
  → VAT ВКЛЮЧАЕТСЯ в счет + комиссия

- EU заказы (Сценарий B): Country != 'GB' и заполнен UK Export Date
  → VAT НЕ включается в счет (возврат "мгновенный"), только комиссия

- UK возвраты (Сценарий C): Country = 'GB' и заполнен UK Export Date
  → VAT ВЫЧИТАЕТСЯ из счета
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# КОНСТАНТЫ
# ============================================================================

# Ставки налогов (можно изменить при необходимости)
UK_VAT_RATE = 0.20  # 20% UK VAT
UK_DUTY_RATE = 0.00  # 0% для парфюмерии (commodity code 3303.00.xxxx)

# Комиссия Duty Refunds
BASE_FEE_PER_ORDER = 1.00  # £1 за каждый заказ
PERCENTAGE_FEE = 0.05  # 5% от (Duty + VAT)

# Курс обмена CAD -> GBP
CAD_TO_GBP_RATE = 0.54

# ============================================================================
# ФУНКЦИИ ДЛЯ ЗАГРУЗКИ И ОБРАБОТКИ ДАННЫХ
# ============================================================================

def load_orders_data(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='Orders', header=1)
        return df
    except Exception as e:
        print(f"✗ Ошибка при загрузке файла: {e}")
        return None


def clean_data(df):
    # Удаляем строки заголовков (если есть)
    df = df[df['FedEx Tracking #'] != 'FedEx Tracking #'].copy()
    
    # Конвертируем типы данных
    df['Line Item Quantity Imported'] = pd.to_numeric(df['Line Item Quantity Imported'], errors='coerce')
    df['Line Item Unit Price CAD'] = pd.to_numeric(df['Line Item Unit Price CAD'], errors='coerce')
    df['Line Item Quantity Exported'] = pd.to_numeric(df['Line Item Quantity Exported'], errors='coerce')
    
    # Конвертируем даты
    df['UK Export Date'] = pd.to_datetime(df['UK Export Date'], errors='coerce')
    
    # Заполняем пропущенные значения
    df['UK Export Date'].fillna(pd.NaT, inplace=True)
    df['Line Item Quantity Exported'].fillna(0, inplace=True)

    return df


def calculate_vat_and_duty(df):

    # Конвертируем цену из CAD в GBP
    df['Line Item Unit Price GBP'] = df['Line Item Unit Price CAD'] * CAD_TO_GBP_RATE
    
    # Рассчитываем общую стоимость позиции
    df['Line Item Total Value GBP'] = df['Line Item Unit Price GBP'] * df['Line Item Quantity Imported']
    
    # Рассчитываем Duty (для парфюмерии обычно 0%)
    df['UK Duty'] = df['Line Item Total Value GBP'] * UK_DUTY_RATE
    
    # Рассчитываем UK VAT (на товары + duty)
    df['UK VAT'] = (df['Line Item Total Value GBP'] + df['UK Duty']) * UK_VAT_RATE

    return df


def classify_orders(df):

    def get_order_type(row):
        if row['Country'] == 'GB':
            # UK заказ
            if pd.isna(row['UK Export Date']):
                return 'UK Order'
            else:
                return 'UK Return'
        else:
            # EU заказ (или другая страна)
            if pd.notna(row['UK Export Date']):
                return 'EU Order'
            else:
                return 'Unknown'  # Это не должно происходить
    
    df['Order Type'] = df.apply(get_order_type, axis=1)

    return df


# ============================================================================
# ФУНКЦИИ ДЛЯ РАСЧЕТА СЧЕТА
# ============================================================================

def calculate_invoice(df):
    """
    Рассчитывает итоговый счет для Samos.
    
    ВАЖНО:
    - UK заказы: VAT включается в счет (+)
    - EU заказы: VAT НЕ включается в счет (возврат мгновенный), только комиссия
    - UK возвраты: VAT вычитается из счета (-)

    Args:
        df (pd.DataFrame): DataFrame с обработанными данными
        
    Returns:
        dict: Словарь с детализацией счета
    """
    invoice = {}
    
    # 1. UK VAT для UK заказов (добавляется к счету)
    uk_orders = df[df['Order Type'] == 'UK Order']
    uk_vat_charged = uk_orders['UK VAT'].sum()
    uk_duty_charged = uk_orders['UK Duty'].sum()
    invoice['UK VAT Charged'] = uk_vat_charged
    invoice['UK Duty Charged'] = uk_duty_charged

    # 2. UK VAT для возвратов (вычитается из счета)
    uk_returns = df[df['Order Type'] == 'UK Return']
    uk_vat_returned = uk_returns['UK VAT'].sum()
    uk_duty_returned = uk_returns['UK Duty'].sum()
    invoice['UK VAT Returned'] = uk_vat_returned
    invoice['UK Duty Returned'] = uk_duty_returned

    # 3. EU заказы - VAT НЕ включается в счет (возврат мгновенный)
    eu_orders = df[df['Order Type'] == 'EU Order']
    eu_vat_not_charged = eu_orders['UK VAT'].sum()
    eu_duty_not_charged = eu_orders['UK Duty'].sum()
    invoice['EU VAT (not charged)'] = eu_vat_not_charged
    invoice['EU Duty (not charged)'] = eu_duty_not_charged

    # 4. Количество заказов (по Parcel ID)
    uk_order_count = uk_orders['Parcel ID'].nunique()
    eu_order_count = eu_orders['Parcel ID'].nunique()
    uk_return_count = uk_returns['Parcel ID'].nunique()
    total_order_count = uk_order_count + eu_order_count

    invoice['UK Order Count'] = uk_order_count
    invoice['EU Order Count'] = eu_order_count
    invoice['UK Return Count'] = uk_return_count
    invoice['Total Order Count'] = total_order_count
    
    # 5. Рассчитываем комиссию Duty Refunds
    # Для UK заказов: £1 + 5% * (Duty + VAT)
    uk_orders_grouped = uk_orders.groupby('Parcel ID').agg({
        'UK Duty': 'sum',
        'UK VAT': 'sum'
    }).reset_index()
    
    uk_orders_grouped['Fee'] = BASE_FEE_PER_ORDER + PERCENTAGE_FEE * (uk_orders_grouped['UK Duty'] + uk_orders_grouped['UK VAT'])
    uk_fee = uk_orders_grouped['Fee'].sum()
    
    # Для EU заказов: £1 + 5% * (Duty + VAT при импорте)
    # ВАЖНО: комиссия берется с VAT, но сам VAT не включается в счет
    eu_orders_grouped = eu_orders.groupby('Parcel ID').agg({
        'UK Duty': 'sum',
        'UK VAT': 'sum'
    }).reset_index()
    
    eu_orders_grouped['Fee'] = BASE_FEE_PER_ORDER + PERCENTAGE_FEE * (eu_orders_grouped['UK Duty'] + eu_orders_grouped['UK VAT'])
    eu_fee = eu_orders_grouped['Fee'].sum()
    
    # Для UK возвратов: комиссия не взимается (или можно взимать, уточните)
    # Пока предполагаем, что комиссия не взимается за возвраты

    total_fee = uk_fee + eu_fee
    
    invoice['Duty Refunds Fee UK'] = uk_fee
    invoice['Duty Refunds Fee EU'] = eu_fee
    invoice['Total Duty Refunds Fee'] = total_fee
    
    # 6. ИТОГО счет
    # ФОРМУЛА: UK VAT (+) - UK VAT возвраты (-) + UK Duty (+) - UK Duty возвраты (-) + Комиссия
    # EU VAT НЕ включается!
    invoice['TOTAL INVOICE'] = (
        uk_vat_charged
        - uk_vat_returned
        + uk_duty_charged
        - uk_duty_returned
        + total_fee
    )

    return invoice


# ============================================================================
# ФУНКЦИИ ДЛЯ ГЕНЕРАЦИИ ОТЧЕТОВ
# ============================================================================

def print_invoice_summary(invoice):
    """
    Выводит сводку по счету.
    
    Args:
        invoice (dict): Словарь с детализацией счета
    """
    print("\n" + "="*80)
    print("СЧЕТ ДЛЯ SAMOS/DECANTBUY")
    print("="*80)
    print(f"\nДата выставления счета: {datetime.now().strftime('%Y-%m-%d')}")
    print(f"Период: [УКАЖИТЕ ПЕРИОД]")
    print(f"\n{'-'*80}")
    print("ДЕТАЛИЗАЦИЯ:")
    print(f"{'-'*80}")
    print(f"\nСЦЕНАРИЙ A: UK ЗАКАЗЫ (товар остается в UK)")
    print(f"  Количество заказов: {invoice['UK Order Count']} шт")
    print(f"  [+] UK VAT (включается в счет):      £{invoice['UK VAT Charged']:>12,.2f}")
    print(f"  [+] UK Duty (парфюмерия 0%):         £{invoice['UK Duty Charged']:>12,.2f}")

    print(f"\nСЦЕНАРИЙ B: EU ЗАКАЗЫ (транзит через UK)")
    print(f"  Количество заказов: {invoice['EU Order Count']} шт")
    print(f"  [ ] UK VAT (НЕ включается в счет):   £{invoice['EU VAT (not charged)']:>12,.2f}")
    print(f"  [ ] UK Duty (НЕ включается):         £{invoice['EU Duty (not charged)']:>12,.2f}")
    print(f"  ℹ️  Возврат VAT происходит мгновенно - не начисляется на Samos")

    if invoice['UK Return Count'] > 0:
        print(f"\nСЦЕНАРИЙ C: UK ВОЗВРАТЫ")
        print(f"  Количество возвратов: {invoice['UK Return Count']} шт")
        print(f"  [-] UK VAT (вычитается из счета):    £{-invoice['UK VAT Returned']:>12,.2f}")
        print(f"  [-] UK Duty (вычитается):            £{-invoice['UK Duty Returned']:>12,.2f}")

    print(f"\n{'-'*80}")
    print("КОМИССИЯ DUTY REFUNDS:")
    print(f"{'-'*80}")
    print(f"UK заказов: {invoice['UK Order Count']:>3} шт × (£1 + 5% VAT+Duty) = £{invoice['Duty Refunds Fee UK']:>12,.2f}")
    print(f"EU заказов: {invoice['EU Order Count']:>3} шт × (£1 + 5% VAT+Duty) = £{invoice['Duty Refunds Fee EU']:>12,.2f}")
    print(f"                                         ─────────────────")
    print(f"Итого комиссия:                          £{invoice['Total Duty Refunds Fee']:>12,.2f}")
    print(f"\n{'='*80}")
    print(f"ИТОГО К ОПЛАТЕ:                          £{invoice['TOTAL INVOICE']:>12,.2f}")
    print(f"{'='*80}")
    print(f"\nУсловия оплаты: 14 дней с даты выставления счета")
    print(f"\nПримечание: ")
    print(f"  • Курс CAD/GBP: {CAD_TO_GBP_RATE}")
    print(f"  • UK VAT: {UK_VAT_RATE*100}%")
    print(f"  • UK Duty (парфюмерия): {UK_DUTY_RATE*100}%")
    print(f"  • EU заказы: VAT не включен в счет (возврат мгновенный)")


def generate_detailed_report(df, output_file='samos_invoice_detail.csv'):
    """
    Генерирует детальный отчет в CSV формате.
    
    Args:
        df (pd.DataFrame): DataFrame с обработанными данными
        output_file (str): Имя выходного файла
    """
    # Группируем по заказам
    report = df.groupby(['Parcel ID', 'Order Type', 'Country']).agg({
        'FedEx Tracking #': 'first',
        'UK Export Date': 'first',
        'UK Export AWB': 'first',
        'Line Item Total Value GBP': 'sum',
        'UK Duty': 'sum',
        'UK VAT': 'sum'
    }).reset_index()
    
    # Сохраняем в CSV
    report.to_csv(output_file, index=False, encoding='utf-8-sig')
    print(f"\n✓ Детальный отчет сохранен в файл: {output_file}")


def generate_order_summary(df, output_file='samos_orders_summary.csv'):
    """
    Генерирует сводку по заказам.
    
    Args:
        df (pd.DataFrame): DataFrame с обработанными данными
        output_file (str): Имя выходного файла
    """
    summary = df.groupby('Order Type').agg({
        'Parcel ID': 'nunique',
        'Line Item Total Value GBP': 'sum',
        'UK Duty': 'sum',
        'UK VAT': 'sum'
    }).reset_index()
    
    summary.columns = ['Order Type', 'Order Count', 'Total Value GBP', 'Total UK Duty', 'Total UK VAT']
    
    # Сохраняем в CSV
    summary.to_csv(output_file, index=False, encoding='utf-8-sig')
    print(f"✓ Сводка по заказам сохранена в файл: {output_file}")


def generate_excel_invoice(df, invoice, output_file='Samos_Invoice_Detailed.xlsx'):
    """
    Генерирует детальный счет в Excel формате с несколькими листами.

    Args:
        df (pd.DataFrame): DataFrame с обработанными данными
        invoice (dict): Словарь с детализацией счета
        output_file (str): Имя выходного файла
    """
    # Создаем Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # ========================================================================
        # ЛИСТ 1: ИТОГОВЫЙ СЧЕТ (INVOICE SUMMARY)
        # ========================================================================
        invoice_data = []
        invoice_data.append(['DUTY REFUNDS LTD', '', '', ''])
        invoice_data.append(['СЧЕТ / INVOICE', '', '', ''])
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['Клиент:', 'SAMOS / DECANTBUY', '', ''])
        invoice_data.append(['Дата:', datetime.now().strftime('%Y-%m-%d'), '', ''])
        invoice_data.append(['Период:', '[УКАЖИТЕ ПЕРИОД]', '', ''])
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['', '', '', ''])

        # Заголовок таблицы
        invoice_data.append(['ОПИСАНИЕ', 'КОЛ-ВО', 'СУММА (£)', 'ПРИМЕЧАНИЕ'])

        # UK заказы
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['СЦЕНАРИЙ A: UK ЗАКАЗЫ', '', '', ''])
        invoice_data.append(['  Количество заказов', invoice['UK Order Count'], '', 'Товар остается в UK'])
        invoice_data.append(['  UK VAT (включается в счет)', '', invoice['UK VAT Charged'], '20% VAT'])
        invoice_data.append(['  UK Duty (парфюмерия)', '', invoice['UK Duty Charged'], '0% Duty'])

        # EU заказы
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['СЦЕНАРИЙ B: EU ЗАКАЗЫ', '', '', ''])
        invoice_data.append(['  Количество заказов', invoice['EU Order Count'], '', 'Транзит через UK'])
        invoice_data.append(['  UK VAT (НЕ включается)', '', invoice['EU VAT (not charged)'], 'Возврат мгновенный'])
        invoice_data.append(['  UK Duty (НЕ включается)', '', invoice['EU Duty (not charged)'], 'Не начисляется'])

        # UK возвраты
        if invoice['UK Return Count'] > 0:
            invoice_data.append(['', '', '', ''])
            invoice_data.append(['СЦЕНАРИЙ C: UK ВОЗВРАТЫ', '', '', ''])
            invoice_data.append(['  Количество возвратов', invoice['UK Return Count'], '', 'Возврат в Канаду'])
            invoice_data.append(['  UK VAT (вычитается)', '', -invoice['UK VAT Returned'], 'Возврат VAT'])
            invoice_data.append(['  UK Duty (вычитается)', '', -invoice['UK Duty Returned'], 'Возврат Duty'])

        # Комиссия
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['КОМИССИЯ DUTY REFUNDS', '', '', ''])
        invoice_data.append(['  UK заказы', invoice['UK Order Count'], invoice['Duty Refunds Fee UK'], '£1 + 5% (Duty+VAT)'])
        invoice_data.append(['  EU заказы', invoice['EU Order Count'], invoice['Duty Refunds Fee EU'], '£1 + 5% (Duty+VAT)'])
        invoice_data.append(['  Итого комиссия', '', invoice['Total Duty Refunds Fee'], ''])

        # Итого
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['ИТОГО К ОПЛАТЕ', '', invoice['TOTAL INVOICE'], ''])
        invoice_data.append(['', '', '', ''])
        invoice_data.append(['Условия оплаты: 14 дней с даты выставления счета', '', '', ''])

        df_invoice = pd.DataFrame(invoice_data)
        df_invoice.to_excel(writer, sheet_name='Invoice Summary', index=False, header=False)

        # ========================================================================
        # ЛИСТ 2: ДЕТАЛИЗАЦИЯ ПО ЗАКАЗАМ С КОМИССИЕЙ
        # ========================================================================

        # Подготовка данных для UK заказов
        uk_orders = df[df['Order Type'] == 'UK Order'].copy()
        uk_orders_detail = uk_orders.groupby('Parcel ID').agg({
            'FedEx Tracking #': 'first',
            'Country': 'first',
            'UK Export Date': 'first',
            'Line Item Total Value GBP': 'sum',
            'UK Duty': 'sum',
            'UK VAT': 'sum'
        }).reset_index()
        uk_orders_detail['Fee'] = BASE_FEE_PER_ORDER + PERCENTAGE_FEE * (uk_orders_detail['UK Duty'] + uk_orders_detail['UK VAT'])
        uk_orders_detail['Order Type'] = 'UK Order'
        uk_orders_detail['Charged to Client'] = uk_orders_detail['UK VAT'] + uk_orders_detail['UK Duty']

        # Подготовка данных для EU заказов
        eu_orders = df[df['Order Type'] == 'EU Order'].copy()
        eu_orders_detail = eu_orders.groupby('Parcel ID').agg({
            'FedEx Tracking #': 'first',
            'Country': 'first',
            'UK Export Date': 'first',
            'Line Item Total Value GBP': 'sum',
            'UK Duty': 'sum',
            'UK VAT': 'sum'
        }).reset_index()
        eu_orders_detail['Fee'] = BASE_FEE_PER_ORDER + PERCENTAGE_FEE * (eu_orders_detail['UK Duty'] + eu_orders_detail['UK VAT'])
        eu_orders_detail['Order Type'] = 'EU Order'
        eu_orders_detail['Charged to Client'] = 0  # VAT не начисляется

        # Подготовка данных для UK возвратов
        uk_returns = df[df['Order Type'] == 'UK Return'].copy()
        if len(uk_returns) > 0:
            uk_returns_detail = uk_returns.groupby('Parcel ID').agg({
                'FedEx Tracking #': 'first',
                'Country': 'first',
                'UK Export Date': 'first',
                'Line Item Total Value GBP': 'sum',
                'UK Duty': 'sum',
                'UK VAT': 'sum'
            }).reset_index()
            uk_returns_detail['Fee'] = 0  # Комиссия за возвраты не взимается
            uk_returns_detail['Order Type'] = 'UK Return'
            uk_returns_detail['Charged to Client'] = -(uk_returns_detail['UK VAT'] + uk_returns_detail['UK Duty'])
        else:
            uk_returns_detail = pd.DataFrame()

        # Объединяем все заказы
        all_orders_detail = pd.concat([uk_orders_detail, eu_orders_detail, uk_returns_detail], ignore_index=True)

        # Переименовываем колонки для отчета
        all_orders_detail = all_orders_detail[[
            'Parcel ID', 'FedEx Tracking #', 'Order Type', 'Country',
            'UK Export Date', 'Line Item Total Value GBP', 'UK Duty', 'UK VAT',
            'Charged to Client', 'Fee'
        ]]
        all_orders_detail.columns = [
            'Parcel ID', 'FedEx Tracking #', 'Order Type', 'Country',
            'UK Export Date', 'Goods Value (£)', 'UK Duty (£)', 'UK VAT (£)',
            'Charged to Client (£)', 'Duty Refunds Fee (£)'
        ]

        all_orders_detail.to_excel(writer, sheet_name='Order Details', index=False)

        # ========================================================================
        # ЛИСТ 3: СВОДКА ПО ТИПАМ ЗАКАЗОВ
        # ========================================================================
        summary_data = []
        summary_data.append(['Order Type', 'Order Count', 'Total Goods Value (£)', 'Total UK Duty (£)',
                           'Total UK VAT (£)', 'Charged to Client (£)', 'Total Fee (£)'])

        for order_type in ['UK Order', 'EU Order', 'UK Return']:
            orders = all_orders_detail[all_orders_detail['Order Type'] == order_type]
            if len(orders) > 0:
                summary_data.append([
                    order_type,
                    len(orders),
                    orders['Goods Value (£)'].sum(),
                    orders['UK Duty (£)'].sum(),
                    orders['UK VAT (£)'].sum(),
                    orders['Charged to Client (£)'].sum(),
                    orders['Duty Refunds Fee (£)'].sum()
                ])

        # Итого
        summary_data.append(['', '', '', '', '', '', ''])
        summary_data.append([
            'TOTAL',
            invoice['Total Order Count'],
            '',
            '',
            '',
            invoice['TOTAL INVOICE'] - invoice['Total Duty Refunds Fee'],
            invoice['Total Duty Refunds Fee']
        ])
        summary_data.append([
            'TOTAL INVOICE',
            '',
            '',
            '',
            '',
            invoice['TOTAL INVOICE'],
            ''
        ])

        df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
        df_summary.to_excel(writer, sheet_name='Summary by Type', index=False)

    # Форматируем Excel файл
    _format_excel_invoice(output_file)

    print(f"\n✓ Детальный счет в Excel сохранен: {output_file}")


def _format_excel_invoice(file_path):
    """
    Применяет форматирование к Excel файлу со счетом.

    Args:
        file_path (str): Путь к Excel файлу
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path)

    # Стили
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    title_font = Font(name='Arial', size=16, bold=True)
    subtitle_font = Font(name='Arial', size=12, bold=True)

    total_font = Font(name='Arial', size=12, bold=True)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    currency_format = '£#,##0.00'

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================================================
    # Форматирование листа Invoice Summary
    # ========================================================================
    ws = wb['Invoice Summary']

    # Заголовок компании
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color='366092')
    ws['A2'].font = Font(name='Arial', size=14, bold=True)

    # Ширина колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    # Форматируем заголовок таблицы (строка 9)
    for col in range(1, 5):
        cell = ws.cell(row=9, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Форматируем суммы (колонка C)
    for row in range(10, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = currency_format

    # Итоговая строка
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value and 'ИТОГО К ОПЛАТЕ' in str(cell.value):
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = total_font
                ws.cell(row=row, column=col).fill = total_fill
                ws.cell(row=row, column=col).border = border

    # ========================================================================
    # Форматирование листа Order Details
    # ========================================================================
    ws = wb['Order Details']

    # Заголовки
    for col in range(1, 11):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Ширина колонок
    ws.column_dimensions['A'].width = 15  # Parcel ID
    ws.column_dimensions['B'].width = 20  # FedEx Tracking
    ws.column_dimensions['C'].width = 12  # Order Type
    ws.column_dimensions['D'].width = 10  # Country
    ws.column_dimensions['E'].width = 15  # Export Date
    ws.column_dimensions['F'].width = 15  # Goods Value
    ws.column_dimensions['G'].width = 12  # UK Duty
    ws.column_dimensions['H'].width = 12  # UK VAT
    ws.column_dimensions['I'].width = 18  # Charged to Client
    ws.column_dimensions['J'].width = 18  # Fee

    # Форматируем суммы
    for row in range(2, ws.max_row + 1):
        for col in [6, 7, 8, 9, 10]:  # Колонки с суммами
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = currency_format

    # ========================================================================
    # Форматирование листа Summary by Type
    # ========================================================================
    ws = wb['Summary by Type']

    # Заголовки
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Ширина колонок
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Форматируем суммы
    for row in range(2, ws.max_row + 1):
        for col in [3, 4, 5, 6, 7]:  # Колонки с суммами
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = currency_format

    # Итоговые строки
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value and str(cell.value) in ['TOTAL', 'TOTAL INVOICE']:
            for col in range(1, 8):
                ws.cell(row=row, column=col).font = total_font
                ws.cell(row=row, column=col).fill = total_fill
                ws.cell(row=row, column=col).border = border

    wb.save(file_path)


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main(file_path):
    """
    Главная функция для запуска всего процесса расчета.
    
    Args:
        file_path (str): Путь к Excel файлу с данными
    """
    print("\n" + "="*80)
    print("DUTY REFUNDS - РАСЧЕТ СЧЕТА ДЛЯ SAMOS/DECANTBUY")
    print("="*80)
    
    # 1. Загрузка данных
    print("\n[1/7] Загрузка данных...")
    df = load_orders_data(file_path)
    if df is None:
        return
    
    # 2. Очистка данных
    print("\n[2/7] Очистка данных...")
    df = clean_data(df)
    
    # 3. Расчет VAT и Duty
    print("\n[3/7] Расчет VAT и Duty...")
    df = calculate_vat_and_duty(df)
    
    # 4. Классификация заказов
    print("\n[4/7] Классификация заказов...")
    df = classify_orders(df)
    
    # 5. Расчет счета
    print("\n[5/7] Расчет счета...")
    invoice = calculate_invoice(df)
    
    # 6. Генерация отчетов
    print("\n[6/7] Генерация отчетов...")
    print_invoice_summary(invoice)
    generate_detailed_report(df)
    generate_order_summary(df)
    
    # 7. Генерация Excel счета
    print("\n[7/7] Генерация Excel счета...")
    generate_excel_invoice(df, invoice)

    print("\n" + "="*80)
    print("ГОТОВО!")
    print("="*80)
    print("\nСозданные файлы:")
    print("  • Samos_Invoice_Detailed.xlsx - ДЕТАЛЬНЫЙ СЧЕТ В EXCEL")
    print("  • samos_invoice_detail.csv - детальный отчет по всем заказам")
    print("  • samos_orders_summary.csv - сводка по типам заказов")
    
    return df, invoice


# ============================================================================
# ЗАПУСК СКРИПТА
# ============================================================================

if __name__ == "__main__":
    # ВАЖНО: Укажите правильный путь к вашему Excel файлу!
    FILE_PATH = 'Copy of UK_drawback_sheet_template.xlsx'
    
    # Запускаем анализ
    df, invoice = main(FILE_PATH)
    
    # Дополнительный анализ (если нужно)
    # Например, можно посмотреть первые строки обработанных данных:
    # print("\n" + "="*80)
    # print("ПЕРВЫЕ 10 СТРОК ОБРАБОТАННЫХ ДАННЫХ:")
    # print("="*80)
    # print(df[['Parcel ID', 'Order Type', 'Country', 'UK VAT', 'UK Duty']].head(10))
